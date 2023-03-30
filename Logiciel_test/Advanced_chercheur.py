import os
import xlwings as xw
import sys
import openpyxl
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import PyQt5.QtGui as QtGui
from PyQt5.QtCore import *

VersionLogiciel = "Beta 1.0" # Version of the program

RealExcelName = "Jauge Pression.xlsx"  # Excel file name
PathToInventaire = "//134.206.175.110/pc2a/Technique/Inventaire Materiel/"  # Path to the inventory folder
# Combine the path and the file name
RealWorkBook = PathToInventaire + RealExcelName
RealSheet = "inventaire jauges de pression" # Sheet name
PathSNF = "//134.206.175.110/pc2a/Technique/Inventaire Materiel/Gaz-Vide-Cryogenie/GAUGES/Données de calibration/" # Path to the Serial number folder

# Chargement du fichier Excel
wb = openpyxl.load_workbook(RealWorkBook)

# Accéder à la feuille "Critères"
ws_criteria = wb[RealSheet]

# Récupérer tous les critères dans la colonne E
criteria_set = set()
for cell in ws_criteria['E']:
    criteria_set.add(str(cell.value))  # convert value to string
criterias = sorted(list(criteria_set))
criterias.remove('gamme')  # remove the first element
criterias.remove('None')  # remove the second element

# Variable en read pour les subwindows
selectedSerialNumber = "Should be overwritten"


def is_available(item):
    # Open the workbook and select the sheet
    workbook = openpyxl.load_workbook(RealWorkBook)
    sheet = workbook[RealSheet]
    # Loop through the rows to find the matching item
    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[1] == item:
            # Check the availability of the item despite the caps
            if row[8] == "stock" or row[8] == "STOCK" or row[8] == "Stock":
                return True
            else:
                return False


# Nouvelle fenetre créer quand on clique sur un objet dans la liste
class AnotherWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Information")
        self.setGeometry(100, 100, 500, 600)
        self.PetitTableau = QTableWidget()
        self.PetitTableau.setRowCount(19)
        self.PetitTableau.setColumnCount(3)


        # Call the function that open the excel file to update itself
        self.Continue = ''
        self.OpenExcelFile()
        if self.Continue:
            print("I continue")
            print(self.Continue)
            # Appel de la fonction qui copie le tableau excel dans le tableau de la fenetre
            self.copyExcelTable()

            self.ClearTheNone()

            # Set the font style and size for the UI elements
            font = QtGui.QFont()
            font.setFamily("Arial")
            font.setPointSize(12)

            # Set the font for the different UI elements
            app.setFont(font)  # Set the font for the application
            self.PetitTableau.setFont(font)  # Set the font for the table widget

            # Set the background color and border for the table widget
            self.PetitTableau.setStyleSheet("QTableWidget {background-color: #F5F5F5; border: 1px solid #CCCCCC;} "
                                            "QHeaderView::section {background-color: #FFFFFF; border: 1px solid #CCCCCC;}")

            # Set the font, background color, and border for the push button
            self.Modifier.setFont(font)  # Set the font for the push button
            self.Modifier.setStyleSheet("QPushButton {background-color: #FFFFFF; border: 1px solid #CCCCCC;}")

            layout = QVBoxLayout()
            layout.addWidget(self.PetitTableau)
            layout.addWidget(self.Modifier)

            self.setLayout(layout)

            self.show()
        else:
            print("I don't continue")
            print(self.Continue)
            self.close()
            return

    def OpenExcelFile(self):
        app = xw.App(visible=False, add_book=False)
        print("Excel app opened")
        if os.path.exists(PathSNF + str(selectedSerialNumber) + ".xlsx"):
            print("File exists")
            self.Continue = True
            pass
        else:
            print("File doesn't exist")
            # pop up a window to warn the user
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Désolé, Maxence n'a pas encore créé le fichier excel \nLes réclamations sont à faire au "
                        "bureau 105")
            msg.setWindowTitle("Attention")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            self.close()
            self.Continue = False
            return

        wb = app.books.open(PathSNF + str(selectedSerialNumber) + ".xlsx")
        print("Workbook opened")
        wb.save()
        print("Workbook saved")
        wb.close()
        print("Workbook closed")
        app.quit()


    def resizeTable(self):
        # Set the width of the columns
        header = self.PetitTableau.horizontalHeader()
        header.setStretchLastSection(False)
        self.PetitTableau.resizeColumnsToContents()

    def ClearTheNone(self):
        # Fill all the empty cells with a blank string
        for row in range(0, 19):
            for col in range(0, 3):
                print(self.PetitTableau.item(row, col).text())
                if self.PetitTableau.item(row, col).text() == "None":
                    self.PetitTableau.setItem(row, col, QTableWidgetItem(""))


    def resizeWindow(self):
        # Get the width of the table
        table_width = self.PetitTableau.horizontalHeader().length()
        table_length = self.PetitTableau.verticalHeader().length()

        # Add some padding to the table width
        padding = 80
        width_with_padding = table_width + padding
        length_with_padding = table_length + padding
        # Resize the window to fit the width of the table
        self.resize(width_with_padding, self.height())
        self.resize(self.width(), length_with_padding)

    # Fonction qui copie un tableau excel dans un tableau de la fenetre
    def copyExcelTable(self):
        print("copyExcelTable")
        # Open the workbook and select the sheet
        workbook = openpyxl.load_workbook(PathSNF + str(selectedSerialNumber) + ".xlsx", data_only=True)
        sheet = workbook["Résumé"]

        # Ajout des titres des colonnes
        self.PetitTableau.setHorizontalHeaderLabels(["Critères", "Valeur"])

        # Copy the Excel table in the new window
        for row in range(1, 20):
            for col in range(1, 4):
                # Copie de la cellule dans le tableau de la fenetre
                self.PetitTableau.setItem(row - 1, col - 1, QTableWidgetItem(str(sheet.cell(row, col).value)))

        # Resize the table and the window
        self.resizeTable()
        self.resizeWindow()
        # self.resizeWindow(self)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.model = QtGui.QStandardItemModel()

        self.setWindowTitle("Trouve ta gauge !")

        self.setGeometry(100, 100, 800, 600)

        layout = QVBoxLayout()

        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)

        # Create a search bar
        qsearch = QLineEdit()
        qsearch.setPlaceholderText("Search by serial number")
        qsearch.textChanged.connect(self.SearchSerialNumber)

        qbox = QComboBox()
        for crit in criterias:
            qbox.addItem(str(crit))
        qbox.addItem("Sélectionner un critère")
        # Make "Sélectionner un critère" the first item in the list
        qbox.insertItem(0, qbox.itemText(qbox.count() - 1)) # Add the last item to the first position (copy it)
        qbox.removeItem(qbox.count() - 1) # Remove the last item

        # Default text in the combo box
        qbox.setCurrentText("Sélectionner un critère")
        # update the list view when the combo box is changed
        qbox.currentIndexChanged.connect(lambda: self.search(criteria=qbox.currentText()))

        self.listView = QListView()
        self.model = QStandardItemModel()
        self.listView.setModel(self.model)
        self.listView.setObjectName("listView-1")
        self.listView.clicked.connect(self.ClickOnObject)

        app.setFont(font)  # Set the font for the application
        qsearch.setFont(font)  # Set the font for the line edit
        qbox.setFont(font)  # Set the font for the combo box

        qsearch.setStyleSheet("QLineEdit {background-color: #F5F5F5; border: 1px solid #CCCCCC;}")
        qbox.setStyleSheet(
            "QComboBox {background-color: #F5F5F5; border: 1px solid #CCCCCC;} QAbstractItemView {background-color: "
            "#FFFFFF; border: 1px solid #CCCCCC;}")

        self.listView.setFont(font)  # Set the font for the list view
        self.listView.setStyleSheet("QListView {background-color: #FFFFFF; border: 1px solid #CCCCCC;}")

        layout.addWidget(qsearch)
        layout.addWidget(qbox)
        layout.addWidget(self.listView)

        widget = QWidget()
        widget.setLayout(layout)

        # Vérification de la version du logiciel
        self.CheckVersion()

        # Set the central widget of the Window. Widget will expand
        # to take up all the space in the window by default.
        self.setCentralWidget(widget)

        # Vérification de la version du logiciel

    def CheckVersion(self):
        wb = openpyxl.load_workbook(RealWorkBook)
        sh_ver = wb['Version']
        version = sh_ver['A2'].value
        if version != VersionLogiciel:
            print("Version du logiciel incompatible")
            QMessageBox.critical(None, "Erreur",
                                 "Version du logiciel incompatible, Merci de télécharger la dernière version")
            sys.exit()

    def SearchSerialNumber(self, serial_number):
        # Open the workbook and select the sheet
        workbook = openpyxl.load_workbook(RealWorkBook)
        sheet = workbook[RealSheet]

        # Clear ListView
        self.model.clear()

        # Loop through the rows to find the matching item even partially
        if len(serial_number) >= 3:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # if the serial number in search bar is in the row, it will add the item to the listview
                if str(serial_number).lower() in str(row[2]).lower():
                    # Add the item to the listview
                    self.model.appendRow(QStandardItem(str(row[1])))
                    # Applying the color function
                    self.color(self.model.item(0))
                else:  # if the serial number is not in the row, it will not add the item to the listview
                    pass
        # if the serial number is less than 3 characters, it will not search anything
        else:
            pass
        # After the loop, it will display "No results" if there is no item in the listview
        if self.model.rowCount() == 0 and len(serial_number) >= 3:
            self.model.appendRow(QStandardItem("No results"))

    # fonction qui choisi une couleur en fonction de la disponibilité
    def color(self, item):
        print("coloring...")
        # Loop through results to find if they are available and display with different colors
        for i in range(0, self.model.rowCount()):
            item = self.model.item(i)
            print("so far so good")
            if is_available(item.text()):
                item.setForeground(QColor(7, 124, 19))
            else:
                item.setForeground(QColor(204, 0, 0))

    def search(self, criteria):
        # Chargement du fichier Excel
        wb = openpyxl.load_workbook(RealWorkBook)
        sheet = wb[RealSheet]
        print(sheet)
        self.model.clear()

        # Loop through the rows to find the matching items
        results = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[4]).lower() == str(criteria).lower():
                self.model.appendRow(QStandardItem(str(row[1])))
                # Application couleur
                self.color(self.model.item(0))
        # If no results, display a message
        if self.model.rowCount() == 0:
            self.model.appendRow(QStandardItem("Aucun résultat"))
        # Return the results
        if str(criteria).lower() == "sélectionner un critère":
            self.model.clear()
        return results

    # Create a new window
    def show_new_window(self, checked):
        self.w = AnotherWindow()
        # self.w.show()

    def AssociateSerialNumber(self, AssociateSerialNumber):
        # Open the workbook and select the sheet
        global selectedSerialNumber
        workbook = openpyxl.load_workbook(RealWorkBook)
        sheet = workbook[RealSheet]

        # Associate the serial number of the clicked item to a variable
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.listView.currentIndex().data():
                AssociateSerialNumber = row[2]
                print("globally selected serial number before", selectedSerialNumber)
                selectedSerialNumber = AssociateSerialNumber
                print("globally selected serial number after", selectedSerialNumber)

    def ClickOnObject(self):
        # Open the workbook and select the sheet
        workbook = openpyxl.load_workbook(RealWorkBook)
        sheet = workbook[RealSheet]
        AssociateSerialNumber = ''

        # Call he function to associate the serial number of the clicked item to a variable
        self.AssociateSerialNumber(AssociateSerialNumber)

        # Open the new window
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == self.listView.currentIndex().data():
                # Create a new window
                self.show_new_window(True)

    def displayObjectsRegardingCriteria(self):
        print("displayObjectsRegardingCriteria")
        criteria = ''

        for child in self.window().children()[1].children():
            if type(child) == QComboBox:
                print(child.currentText())
                criteria = child.currentText()

        objects = self.search(criteria)

        u = 0
        for object in objects:

            children = self.window().children()[1].children()
            print(object)
            print(children)
            self.layout().addWidget(QLabel(object[1]))
            for i in range(u, len(children)):
                if type(children[i]) == QLabel:
                    u = i
                    print(children[i].setText(object[1]))

        return ''


app = QApplication(sys.argv)
window = MainWindow()
window.show()

app.exec()
