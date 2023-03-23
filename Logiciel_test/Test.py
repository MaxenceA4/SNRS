import openpyxl
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk

# Chargement du fichier Excel
wb = openpyxl.load_workbook('D:/Vrai_bureau/Logiciel_test/Inventaire.xlsx')

# Accéder à la feuille "Utilisateurs"
ws_users = wb['Utilisateurs']

# Récupérer tous les noms d'utilisateurs dans la colonne A
users = []
for cell in ws_users['A']:
    users.append(cell.value)

# Accéder à la feuille "Critères"
ws_criteria = wb['Critères']

# Récupérer tous les critères dans la colonne A
criteria_set = set()
for cell in ws_criteria['A']:
    criteria_set.add(cell.value)
criteria = list(criteria_set)

# Créer une fenêtre tkinter
root = tk.Tk()

# Définir le titre de la fenêtre
root.title("Inventaire d'objets")

# Définir la taille de la fenêtre
root.geometry("800x400")

# Définir la couleur de fond de la fenêtre
root.configure(background='#F7D6B2')

# Créer un cadre pour les menus déroulants
frame = tk.Frame(root, bg='#F7D6B2')
frame.pack(pady=20)

# Créer le menu déroulant pour les utilisateurs
user_var = tk.StringVar(root)
user_var.set("Sélectionner un utilisateur")
user_menu = tk.OptionMenu(frame, user_var, *users)
user_menu.config(bg='#F3C29B', width=25, font=('Helvetica', 12))
user_menu.pack(side='left', padx=10)

# Créer le menu déroulant pour les critères
criteria_var = tk.StringVar(root)
criteria_var.set("Sélectionner un critère")
criteria_menu = tk.OptionMenu(frame, criteria_var, *criteria)
criteria_menu.config(bg='#F3C29B', width=25, font=('Helvetica', 12))
criteria_menu.pack(side='right', padx=10)

# Chargement de l'icône de loupe
img = Image.open("D:/Vrai_bureau/Logiciel_test/loupe.png")
img = img.resize((50, 50), Image.ANTIALIAS)  # Redimensionner l'icône
search_icon = ImageTk.PhotoImage(img)

# Ajouter un bouton pour valider la sélection
submit_button = tk.Button(root, text="", image=search_icon, bg='#F3C29B', font=('Helvetica', 12))
submit_button.pack(pady=20)

# Créer une zone de texte pour afficher les résultats de recherche
result_label = tk.Label(root, text="", bg='#F7D6B2', font=('Helvetica', 12))
result_label.pack(pady=20)


# Fonction pour effectuer une recherche

def search():
    # Get the user and criteria
    user = user_var.get()
    criteria = criteria_var.get()

    # Check that both a user and criteria have been selected
    if user == "Sélectionner un utilisateur" or criteria == "Sélectionner un critère":
        error_label.config(text="Veuillez sélectionner à la fois un utilisateur et un critère.")
        return

    # Clear any previous error message
    error_label.config(text="")

    # Open the workbook and select the sheet
    workbook = openpyxl.load_workbook("Inventaire.xlsx")
    sheet = workbook["Liste"]

    # Loop through the rows to find the matching items
    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]).lower() == str(criteria).lower():
            results.append(row[1])

    # Display the results
    if results:
        result_label.config(text="\n".join(results))
    else:
        result_label.config(text="Aucun résultat trouvé pour le critère sélectionné.")




submit_button.config(text="", image=search_icon)
submit_button.config(command=search)

error_label = tk.Label(root, bg='#F7D6B2', fg='red', font=('Helvetica', 12))
error_label.pack(pady=10)

# Démarrer la boucle principale de la fenêtre
root.mainloop()